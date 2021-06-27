from dotenv import load_dotenv
from pathlib import Path
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Color, PatternFill
import os
import argparse
import requests
import urllib.parse
import time

# https://positionstack.com/documentation

COLOR_BLACK = "000000"
COLOR_WHITE = "FFFFFF"
COLOR_RED = "FF5959"
COLOR_ORANGE = "FFDF50"

api_key = None
process_mode = 1    # 1 = only if coordinates are empty
                    # 2 = also process if confidence is below 1

def get_coordinates(address:str, zipcode:str, city:str, country_code:str):

    url = "http://api.positionstack.com/v1/forward"
    url += f"?access_key={api_key}"
    q = f"{address}, {zipcode} {city}"
    url += f"&query={urllib.parse.quote(q)}"
    url += f"&country={urllib.parse.quote(country_code)}"
    response = requests.get(url)
    if response.status_code == 200:
        js = response.json()
        if "data" in js:
            data = js["data"]
            if len(data) >= 1:
                rec1 = data[0]
                # print(rec1)
                if rec1 == []:
                    return None
                return rec1
    else:
        print(f"Error {response.status_code} : {url}")

    return None

def add_cell(sheet, column, row, value, bold = False):
    char = get_column_letter(column)
    c = sheet[char + str(row)]
    c.value = value
    if bold:
        c.font = Font(bold = True)

def set_cell_color(sheet, column, row, color):
    char = get_column_letter(column)
    c = sheet[char + str(row)]
    c.fill = PatternFill(patternType = "solid", fgColor = color)

def process(infile:str):
    print("Reading...")
    workbook = load_workbook(infile)
    worksheet = workbook.active

    add_cell(worksheet, 9, 1, "Latitude", bold = True)
    add_cell(worksheet, 10, 1, "Longitude", bold = True)
    add_cell(worksheet, 11, 1, "Confidence", bold = True)
    add_cell(worksheet, 12, 1, "Region", bold = True)
    add_cell(worksheet, 13, 1, "Region_Code", bold = True)
    add_cell(worksheet, 14, 1, "County", bold = True)
    add_cell(worksheet, 15, 1, "Locality", bold = True)
    add_cell(worksheet, 16, 1, "Area", bold = True)
    add_cell(worksheet, 17, 1, "Neighbourhood", bold = True)
    add_cell(worksheet, 18, 1, "Country", bold = True)
    add_cell(worksheet, 19, 1, "Country Code", bold = True)
    add_cell(worksheet, 20, 1, "Continent", bold = True)
    add_cell(worksheet, 21, 1, "Label", bold = True)

    print("Processing...")
    line = 1
    changed = False
    while True:
        line += 1
        name = worksheet[f"A{line}"].value
        if name == "" or name is None:
            break
        address = worksheet[f"B{line}"].value
        zipcode = worksheet[f"C{line}"].value
        city = worksheet[f"D{line}"].value
        _ = worksheet[f"E{line}"].value
        _ = worksheet[f"F{line}"].value
        _ = worksheet[f"G{line}"].value
        country_code = worksheet[f"H{line}"].value

        latitude = worksheet[f"I{line}"].value
        longitude = worksheet[f"J{line}"].value
        confidence = worksheet[f"K{line}"].value

        if process_mode == 1:
            if latitude is None or longitude is None:
                pass
            else:
                continue
        else:
            if latitude is None or longitude is None or confidence != 1:
                pass
            else:
                continue

        try:
            changed = True
            coords = get_coordinates(address, zipcode, city, country_code)
            if coords is not None:
                latitude = coords["latitude"]
                longitude = coords["longitude"]
                confidence = coords["confidence"]

                region = coords["region"]
                region_code = coords["region_code"]
                county = coords["county"]
                locality = coords["locality"]
                administrative_area = coords["administrative_area"]
                neighbourhood = coords["neighbourhood"]
                country = coords["country"]
                country_code = coords["country_code"]
                continent = coords["continent"]
                label = coords["label"]

                print(f"{name} => {latitude},{longitude}")

                add_cell(worksheet, 9, line, latitude)
                add_cell(worksheet, 10, line, longitude)
                add_cell(worksheet, 11, line, confidence)
                add_cell(worksheet, 12, line, region)
                add_cell(worksheet, 13, line, region_code)
                add_cell(worksheet, 14, line, country)
                add_cell(worksheet, 15, line, locality)
                add_cell(worksheet, 16, line, administrative_area)
                add_cell(worksheet, 17, line, neighbourhood)
                add_cell(worksheet, 18, line, country)
                add_cell(worksheet, 19, line, country_code)
                add_cell(worksheet, 20, line, continent)
                add_cell(worksheet, 21, line, label)

                if confidence == 1:
                    set_cell_color(worksheet, 1, line, COLOR_WHITE)
                    set_cell_color(worksheet, 9, line, COLOR_WHITE)
                    set_cell_color(worksheet, 10, line, COLOR_WHITE)
                    set_cell_color(worksheet, 11, line, COLOR_WHITE)
                else:
                    set_cell_color(worksheet, 1, line, COLOR_ORANGE)
                    set_cell_color(worksheet, 9, line, COLOR_ORANGE)
                    set_cell_color(worksheet, 10, line, COLOR_ORANGE)
                    set_cell_color(worksheet, 11, line, COLOR_ORANGE)
            else:
                print(f"{name} => Not found")
                for colnum in range(9, 22):
                    add_cell(worksheet, colnum, line, "")
                    set_cell_color(worksheet, 1, line, COLOR_RED)

        except Exception as e:
            print(f"Error: Line {line} - {name}")
            print(e)
        else:
            pass
        finally:
            pass

        time.sleep(0.2)

    if changed:
        print("Writing...")
        workbook.save(infile)
    else:
        print("No changes")

def main():
    global api_key, process_mode
    load_dotenv()
    env_path = Path('.')
    env_file = '.env'
    load_dotenv(dotenv_path = os.path.join(env_path, env_file))
    api_key = os.getenv("API_KEY")

    parser = argparse.ArgumentParser()
    parser.add_argument("inputfile", help = "Name of the Excel file with addresses")
    parser.add_argument("-u", "--update", action = "store_true", help = "Update the rows where longitude or latitude are empty. (the default)")
    parser.add_argument("-r", "--redo", action = "store_true", help = "Update the rows where longitude or latitude are empty, or where the confidence is not equal to 1.")
    args = parser.parse_args()

    if args.update:
        process_mode = 1
    if args.redo:
        process_mode = 2

    if os.path.isfile(args.inputfile) == False:
        print(f"Error: file not found: {args.inputfile}")
        return

    infile = args.inputfile

    if infile[-5:].lower() != ".xlsx":
        print("Error: inputfile must be an .xlsx file")
        return

    process(infile)

if __name__ == "__main__":
    main()

